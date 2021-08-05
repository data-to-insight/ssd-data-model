from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from rtofdata.config import output_dir, output_filename_base
from rtofdata.spec_parser import Specification


def write_excel_specification(spec: Specification):
    write_dimensions(spec)
    write_records(spec)


def write_records(spec: Specification):
    wb = Workbook()

    for flow in spec.records_by_flow:
        r = flow.record
        ws = wb.create_sheet(r.id)
        ws.append((f"{r.id}", ))
        ws.append((f"{r.description}", ))
        ws.append([])

        data = [[f.id, f.name, f.primary_key, f.type, f.dimensions.id if f.dimensions else ""] for f in r.fields]
        if len(data) == 0:
            data = [["No Data Available", ""]]

        ws.append(["ID", "Name", "Primary Key", "Type", "Dimension"])
        for row in data:
            ws.append(row)

        start_row = 4
        tab = Table(displayName=r.id, ref=f"A{start_row}:E{start_row+len(data)}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        ws.add_table(tab)
        for c in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[c].width = 45
        ws['A1'].style = "Headline 1"

    ws = wb.worksheets[0]
    wb.remove_sheet(ws)
    wb.save(output_dir / f"{output_filename_base}-records.xlsx" )


def write_dimensions(spec: Specification):
    wb = Workbook()

    for d in spec.dimensions:
        ws = wb.create_sheet(d.id)

        data = [[v.value, v.description] for v in d.dimensions]
        if len(data) == 0:
            data = [["No Data Available", ""]]

        ws.append((d.id, ))
        ws.append([])

        ws.append(["Value", "Description"])
        for row in data:
            ws.append(row)

        start_row = 3
        tab = Table(displayName=d.id, ref=f"A{start_row}:B{start_row + len(data)}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        ws.add_table(tab)
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 45
        ws['A1'].style = "Headline 1"

    ws = wb.worksheets[0]
    wb.remove_sheet(ws)
    wb.save(output_dir / f"{output_filename_base}-dimensions.xlsx" )

